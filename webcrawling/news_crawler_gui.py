#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
PyQt5 기반 GUI 뉴스 크롤링 프로그램
사용자 친화적인 인터페이스로 뉴스 크롤링 및 저장 기능 제공
"""

import sys
import os
from PyQt5.QtWidgets import (QApplication, QMainWindow, QVBoxLayout, QHBoxLayout, 
                            QWidget, QPushButton, QTextEdit, QLabel, QLineEdit, 
                            QComboBox, QProgressBar, QGroupBox, QCheckBox, 
                            QFileDialog, QMessageBox, QSplitter, QTabWidget,
                            QTableWidget, QTableWidgetItem, QHeaderView)
from PyQt5.QtCore import QThread, pyqtSignal, Qt, QTimer
from PyQt5.QtGui import QFont, QIcon, QPalette, QColor
import requests
from bs4 import BeautifulSoup
import time
import csv
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment


class CrawlerThread(QThread):
    """백그라운드에서 크롤링을 수행하는 워커 스레드"""
    
    # 시그널 정의
    progress_updated = pyqtSignal(int)  # 진행률
    status_updated = pyqtSignal(str)    # 상태 메시지
    result_ready = pyqtSignal(list)     # 크롤링 결과
    error_occurred = pyqtSignal(str)    # 에러 메시지
    
    def __init__(self, keyword, sources):
        super().__init__()
        self.keyword = keyword
        self.sources = sources
        self.session = requests.Session()
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        })
        self.session.verify = False
        # SSL 경고 메시지 숨기기
        import urllib3
        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
    
    def run(self):
        """크롤링 실행"""
        try:
            all_titles = []
            total_sources = len(self.sources)
            
            for idx, source in enumerate(self.sources):
                self.status_updated.emit(f"{source}에서 크롤링 중...")
                
                if source == "구글 뉴스":
                    titles = self.crawl_google_news()
                elif source == "다음 뉴스":
                    titles = self.crawl_daum_news()
                elif source == "연합뉴스":
                    titles = self.crawl_yna_news()
                else:
                    titles = []
                
                # 중복 제거하여 추가
                for title in titles:
                    if title not in all_titles:
                        all_titles.append(title)
                
                # 진행률 업데이트
                progress = int((idx + 1) / total_sources * 100)
                self.progress_updated.emit(progress)
                
                # 요청 간격 조절
                time.sleep(1)
            
            self.status_updated.emit(f"크롤링 완료! 총 {len(all_titles)}개의 뉴스 제목 수집")
            self.result_ready.emit(all_titles)
            
        except Exception as e:
            self.error_occurred.emit(f"크롤링 중 오류 발생: {str(e)}")
    
    def crawl_google_news(self):
        """구글 뉴스 크롤링"""
        try:
            url = f"https://news.google.com/rss/search?q={self.keyword}&hl=ko&gl=KR&ceid=KR:ko"
            response = self.session.get(url)
            response.raise_for_status()
            response.encoding = 'utf-8'
            
            soup = BeautifulSoup(response.text, 'xml')
            titles = []
            
            items = soup.find_all('item')
            for item in items:
                title_element = item.find('title')
                if title_element:
                    title = title_element.get_text().strip()
                    if ' - ' in title:
                        title = title.split(' - ')[0]
                    if title and len(title) > 5:
                        titles.append(title)
            
            return titles[:50]  # 최대 50개만 반환
            
        except Exception as e:
            self.status_updated.emit(f"구글 뉴스 크롤링 실패: {str(e)}")
            return []
    
    def crawl_daum_news(self):
        """다음 뉴스 크롤링"""
        try:
            url = f"https://search.daum.net/search?w=news&q={self.keyword}&DA=PGD&spacing=0"
            response = self.session.get(url)
            response.raise_for_status()
            response.encoding = 'utf-8'
            
            soup = BeautifulSoup(response.text, 'html.parser')
            titles = []
            
            title_elements = soup.find_all('a', class_='f_link_b')
            for element in title_elements:
                title = element.get_text().strip()
                if title and len(title) > 5:
                    titles.append(title)
            
            return titles[:30]  # 최대 30개만 반환
            
        except Exception as e:
            self.status_updated.emit(f"다음 뉴스 크롤링 실패: {str(e)}")
            return []
    
    def crawl_yna_news(self):
        """연합뉴스 크롤링"""
        try:
            url = f"https://www.yna.co.kr/search/index?query={self.keyword}"
            response = self.session.get(url)
            response.raise_for_status()
            response.encoding = 'utf-8'
            
            soup = BeautifulSoup(response.text, 'html.parser')
            titles = []
            
            title_elements = soup.find_all('strong', class_='tit-news')
            for element in title_elements:
                link = element.find('a')
                if link:
                    title = link.get_text().strip()
                    if title and len(title) > 5:
                        titles.append(title)
            
            return titles[:30]  # 최대 30개만 반환
            
        except Exception as e:
            self.status_updated.emit(f"연합뉴스 크롤링 실패: {str(e)}")
            return []


class NewsCrawlerGUI(QMainWindow):
    """메인 GUI 클래스"""
    
    def __init__(self):
        super().__init__()
        self.news_data = []  # 크롤링된 뉴스 데이터 저장
        self.init_ui()
        
    def init_ui(self):
        """UI 초기화"""
        self.setWindowTitle("📰 뉴스 크롤링 프로그램 v2.0")
        self.setGeometry(100, 100, 1000, 700)
        
        # 중앙 위젯 설정
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        # 메인 레이아웃
        main_layout = QVBoxLayout(central_widget)
        
        # 타이틀
        title_label = QLabel("🚀 AI 뉴스 크롤링 도구")
        title_font = QFont("Arial", 16, QFont.Bold)
        title_label.setFont(title_font)
        title_label.setAlignment(Qt.AlignCenter)
        title_label.setStyleSheet("color: #2c3e50; margin: 10px;")
        main_layout.addWidget(title_label)
        
        # 스플리터로 상하 분할
        splitter = QSplitter(Qt.Vertical)
        main_layout.addWidget(splitter)
        
        # 상단: 설정 패널
        settings_widget = self.create_settings_panel()
        splitter.addWidget(settings_widget)
        
        # 하단: 결과 패널
        results_widget = self.create_results_panel()
        splitter.addWidget(results_widget)
        
        # 스플리터 비율 설정
        splitter.setStretchFactor(0, 1)
        splitter.setStretchFactor(1, 3)
        
        # 스타일 적용
        self.apply_styles()
    
    def create_settings_panel(self):
        """설정 패널 생성"""
        settings_group = QGroupBox("🔧 크롤링 설정")
        layout = QVBoxLayout(settings_group)
        
        # 키워드 입력
        keyword_layout = QHBoxLayout()
        keyword_layout.addWidget(QLabel("검색 키워드:"))
        self.keyword_input = QLineEdit()
        self.keyword_input.setText("반도체 관련주")
        self.keyword_input.setPlaceholderText("크롤링할 키워드를 입력하세요")
        keyword_layout.addWidget(self.keyword_input)
        layout.addLayout(keyword_layout)
        
        # 소스 선택
        sources_layout = QHBoxLayout()
        sources_layout.addWidget(QLabel("크롤링 소스:"))
        
        self.google_checkbox = QCheckBox("구글 뉴스")
        self.google_checkbox.setChecked(True)
        self.daum_checkbox = QCheckBox("다음 뉴스")
        self.yna_checkbox = QCheckBox("연합뉴스")
        
        sources_layout.addWidget(self.google_checkbox)
        sources_layout.addWidget(self.daum_checkbox)
        sources_layout.addWidget(self.yna_checkbox)
        sources_layout.addStretch()
        layout.addLayout(sources_layout)
        
        # 컨트롤 버튼들
        buttons_layout = QHBoxLayout()
        
        self.start_button = QPushButton("🚀 크롤링 시작")
        self.start_button.clicked.connect(self.start_crawling)
        
        self.stop_button = QPushButton("⏹️ 중지")
        self.stop_button.clicked.connect(self.stop_crawling)
        self.stop_button.setEnabled(False)
        
        self.save_csv_button = QPushButton("📄 CSV 저장")
        self.save_csv_button.clicked.connect(self.save_to_csv)
        self.save_csv_button.setEnabled(False)
        
        self.save_excel_button = QPushButton("📊 Excel 저장")
        self.save_excel_button.clicked.connect(self.save_to_excel)
        self.save_excel_button.setEnabled(False)
        
        buttons_layout.addWidget(self.start_button)
        buttons_layout.addWidget(self.stop_button)
        buttons_layout.addWidget(self.save_csv_button)
        buttons_layout.addWidget(self.save_excel_button)
        buttons_layout.addStretch()
        
        layout.addLayout(buttons_layout)
        
        # 진행률 표시
        progress_layout = QVBoxLayout()
        self.status_label = QLabel("준비됨")
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        
        progress_layout.addWidget(self.status_label)
        progress_layout.addWidget(self.progress_bar)
        layout.addLayout(progress_layout)
        
        return settings_group
    
    def create_results_panel(self):
        """결과 패널 생성"""
        results_group = QGroupBox("📋 크롤링 결과")
        layout = QVBoxLayout(results_group)
        
        # 탭 위젯 생성
        tab_widget = QTabWidget()
        
        # 테이블 뷰 탭
        self.table_widget = QTableWidget()
        self.table_widget.setColumnCount(3)
        self.table_widget.setHorizontalHeaderLabels(["번호", "뉴스 제목", "수집 시간"])
        
        # 테이블 컬럼 너비 설정
        header = self.table_widget.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.Fixed)
        header.setSectionResizeMode(1, QHeaderView.Stretch)
        header.setSectionResizeMode(2, QHeaderView.Fixed)
        self.table_widget.setColumnWidth(0, 60)
        self.table_widget.setColumnWidth(2, 150)
        
        tab_widget.addTab(self.table_widget, "📊 테이블 뷰")
        
        # 텍스트 뷰 탭
        self.results_text = QTextEdit()
        self.results_text.setReadOnly(True)
        self.results_text.setPlaceholderText("크롤링 결과가 여기에 표시됩니다...")
        tab_widget.addTab(self.results_text, "📝 텍스트 뷰")
        
        layout.addWidget(tab_widget)
        
        # 통계 정보
        stats_layout = QHBoxLayout()
        self.stats_label = QLabel("수집된 뉴스: 0개")
        stats_layout.addWidget(self.stats_label)
        stats_layout.addStretch()
        layout.addLayout(stats_layout)
        
        return results_group
    
    def apply_styles(self):
        """스타일 적용"""
        style = """
        QMainWindow {
            background-color: #f8f9fa;
        }
        QGroupBox {
            font-weight: bold;
            border: 2px solid #3498db;
            border-radius: 8px;
            margin: 5px;
            padding-top: 10px;
        }
        QGroupBox::title {
            subcontrol-origin: margin;
            left: 10px;
            padding: 0 5px 0 5px;
            color: #2c3e50;
        }
        QPushButton {
            background-color: #3498db;
            color: white;
            border: none;
            padding: 8px 16px;
            border-radius: 4px;
            font-weight: bold;
        }
        QPushButton:hover {
            background-color: #2980b9;
        }
        QPushButton:pressed {
            background-color: #21618c;
        }
        QPushButton:disabled {
            background-color: #bdc3c7;
            color: #7f8c8d;
        }
        QLineEdit {
            padding: 8px;
            border: 2px solid #bdc3c7;
            border-radius: 4px;
            font-size: 12px;
        }
        QLineEdit:focus {
            border-color: #3498db;
        }
        QTextEdit {
            border: 2px solid #bdc3c7;
            border-radius: 4px;
            font-family: Consolas, monospace;
        }
        QTableWidget {
            gridline-color: #bdc3c7;
            border: 2px solid #bdc3c7;
            border-radius: 4px;
        }
        QTableWidget::item {
            padding: 8px;
        }
        QProgressBar {
            border: 2px solid #bdc3c7;
            border-radius: 4px;
            text-align: center;
        }
        QProgressBar::chunk {
            background-color: #27ae60;
            border-radius: 2px;
        }
        """
        self.setStyleSheet(style)
    
    def start_crawling(self):
        """크롤링 시작"""
        keyword = self.keyword_input.text().strip()
        if not keyword:
            QMessageBox.warning(self, "경고", "키워드를 입력해주세요!")
            return
        
        # 선택된 소스 확인
        sources = []
        if self.google_checkbox.isChecked():
            sources.append("구글 뉴스")
        if self.daum_checkbox.isChecked():
            sources.append("다음 뉴스")
        if self.yna_checkbox.isChecked():
            sources.append("연합뉴스")
        
        if not sources:
            QMessageBox.warning(self, "경고", "최소 하나의 소스를 선택해주세요!")
            return
        
        # UI 상태 변경
        self.start_button.setEnabled(False)
        self.stop_button.setEnabled(True)
        self.save_csv_button.setEnabled(False)
        self.save_excel_button.setEnabled(False)
        self.progress_bar.setVisible(True)
        self.progress_bar.setValue(0)
        
        # 기존 결과 초기화
        self.news_data.clear()
        self.table_widget.setRowCount(0)
        self.results_text.clear()
        
        # 크롤링 스레드 시작
        self.crawler_thread = CrawlerThread(keyword, sources)
        self.crawler_thread.progress_updated.connect(self.update_progress)
        self.crawler_thread.status_updated.connect(self.update_status)
        self.crawler_thread.result_ready.connect(self.handle_results)
        self.crawler_thread.error_occurred.connect(self.handle_error)
        self.crawler_thread.start()
    
    def stop_crawling(self):
        """크롤링 중지"""
        if hasattr(self, 'crawler_thread') and self.crawler_thread.isRunning():
            self.crawler_thread.terminate()
            self.crawler_thread.wait()
        
        self.reset_ui_state()
        self.update_status("크롤링이 중지되었습니다.")
    
    def update_progress(self, value):
        """진행률 업데이트"""
        self.progress_bar.setValue(value)
    
    def update_status(self, message):
        """상태 메시지 업데이트"""
        self.status_label.setText(f"상태: {message}")
    
    def handle_results(self, titles):
        """크롤링 결과 처리"""
        self.news_data = titles
        current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        # 테이블 업데이트
        self.table_widget.setRowCount(len(titles))
        for idx, title in enumerate(titles):
            self.table_widget.setItem(idx, 0, QTableWidgetItem(str(idx + 1)))
            self.table_widget.setItem(idx, 1, QTableWidgetItem(title))
            self.table_widget.setItem(idx, 2, QTableWidgetItem(current_time))
        
        # 텍스트 뷰 업데이트
        text_content = f"🔍 검색 키워드: {self.keyword_input.text()}\n"
        text_content += f"📅 수집 시간: {current_time}\n"
        text_content += f"📊 총 {len(titles)}개의 뉴스 제목\n\n"
        text_content += "=" * 60 + "\n"
        
        for idx, title in enumerate(titles, 1):
            text_content += f"{idx:3d}. {title}\n"
        
        self.results_text.setPlainText(text_content)
        
        # 통계 정보 업데이트
        self.stats_label.setText(f"수집된 뉴스: {len(titles)}개")
        
        # UI 상태 복원
        self.reset_ui_state()
        self.save_csv_button.setEnabled(True)
        self.save_excel_button.setEnabled(True)
        
        QMessageBox.information(self, "완료", f"총 {len(titles)}개의 뉴스 제목을 수집했습니다!")
    
    def handle_error(self, error_message):
        """에러 처리"""
        self.reset_ui_state()
        QMessageBox.critical(self, "오류", error_message)
    
    def reset_ui_state(self):
        """UI 상태 초기화"""
        self.start_button.setEnabled(True)
        self.stop_button.setEnabled(False)
        self.progress_bar.setVisible(False)
        self.update_status("준비됨")
    
    def save_to_csv(self):
        """CSV 파일로 저장"""
        if not self.news_data:
            QMessageBox.warning(self, "경고", "저장할 데이터가 없습니다!")
            return
        
        file_path, _ = QFileDialog.getSaveFileName(
            self, "CSV 파일 저장", "news_result.csv", "CSV 파일 (*.csv)"
        )
        
        if file_path:
            try:
                with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
                    writer = csv.writer(csvfile)
                    writer.writerow(['번호', '뉴스 제목', '수집시간'])
                    
                    current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    for idx, title in enumerate(self.news_data, 1):
                        writer.writerow([idx, title, current_time])
                
                QMessageBox.information(self, "저장 완료", f"CSV 파일이 저장되었습니다:\n{file_path}")
                
            except Exception as e:
                QMessageBox.critical(self, "저장 오류", f"CSV 파일 저장 중 오류가 발생했습니다:\n{str(e)}")
    
    def save_to_excel(self):
        """Excel 파일로 저장"""
        if not self.news_data:
            QMessageBox.warning(self, "경고", "저장할 데이터가 없습니다!")
            return
        
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Excel 파일 저장", "news_result.xlsx", "Excel 파일 (*.xlsx)"
        )
        
        if file_path:
            try:
                # 새 워크북 생성
                wb = Workbook()
                ws = wb.active
                ws.title = "뉴스 제목"
                
                # 헤더 스타일 설정
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                center_alignment = Alignment(horizontal='center', vertical='center')
                
                # 헤더 작성
                headers = ['번호', '뉴스 제목', '수집시간']
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = header_border
                    cell.alignment = center_alignment
                
                # 데이터 입력
                current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                for idx, title in enumerate(self.news_data, 1):
                    ws.cell(row=idx+1, column=1, value=idx).alignment = center_alignment
                    ws.cell(row=idx+1, column=2, value=title)
                    ws.cell(row=idx+1, column=3, value=current_time).alignment = center_alignment
                
                # 컬럼 너비 자동 조정
                ws.column_dimensions['A'].width = 8
                ws.column_dimensions['B'].width = 80
                ws.column_dimensions['C'].width = 20
                
                # 데이터 영역 테두리 추가
                data_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                for row in range(2, len(self.news_data) + 2):
                    for col in range(1, 4):
                        ws.cell(row=row, column=col).border = data_border
                
                # 제목 행 고정
                ws.freeze_panes = 'A2'
                
                # 파일 저장
                wb.save(file_path)
                
                QMessageBox.information(self, "저장 완료", f"Excel 파일이 저장되었습니다:\n{file_path}")
                
            except Exception as e:
                QMessageBox.critical(self, "저장 오류", f"Excel 파일 저장 중 오류가 발생했습니다:\n{str(e)}")


def main():
    """메인 함수"""
    app = QApplication(sys.argv)
    
    # 애플리케이션 정보 설정
    app.setApplicationName("뉴스 크롤링 프로그램")
    app.setApplicationVersion("2.0")
    app.setOrganizationName("WebCrawler Corp")
    
    # 한글 폰트 설정
    font = QFont("맑은 고딕", 9)
    app.setFont(font)
    
    # GUI 실행
    window = NewsCrawlerGUI()
    window.show()
    
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()
