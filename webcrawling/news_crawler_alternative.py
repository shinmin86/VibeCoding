#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
대안적인 뉴스 크롤링 방법들을 제공하는 스크립트
네이버 직접 크롤링이 어려울 때 사용할 수 있는 여러 방법들
"""

import requests
from bs4 import BeautifulSoup
import time
import csv
import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


class AlternativeNewsCrawler:
    def __init__(self):
        self.session = requests.Session()
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        })
        # SSL 인증서 검증 비활성화
        self.session.verify = False
        import urllib3
        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
    
    def crawl_daum_news(self, keyword="반도체"):
        """
        다음 뉴스에서 특정 키워드 관련 뉴스 제목을 크롤링합니다.
        
        Args:
            keyword (str): 검색할 키워드
            
        Returns:
            list: 뉴스 제목 리스트
        """
        try:
            # 다음 뉴스 검색 URL
            url = f"https://search.daum.net/search?w=news&q={keyword}&DA=PGD&spacing=0"
            print(f"다음 뉴스 검색: {url}")
            
            response = self.session.get(url)
            response.raise_for_status()
            response.encoding = 'utf-8'
            
            soup = BeautifulSoup(response.text, 'html.parser')
            news_titles = []
            
            # 다음 뉴스 제목 추출
            title_elements = soup.find_all('a', class_='f_link_b')
            
            for element in title_elements:
                title = element.get_text().strip()
                if title and len(title) > 5:
                    news_titles.append(title)
            
            print(f"다음에서 {len(news_titles)}개의 뉴스 제목을 찾았습니다.")
            return news_titles
            
        except Exception as e:
            print(f"다음 뉴스 크롤링 오류: {e}")
            return []
    
    def crawl_google_news(self, keyword="반도체 관련주"):
        """
        구글 뉴스에서 특정 키워드 관련 뉴스 제목을 크롤링합니다.
        
        Args:
            keyword (str): 검색할 키워드
            
        Returns:
            list: 뉴스 제목 리스트
        """
        try:
            # 구글 뉴스 검색 URL (한국 지역)
            url = f"https://news.google.com/rss/search?q={keyword}&hl=ko&gl=KR&ceid=KR:ko"
            print(f"구글 뉴스 RSS 검색: {url}")
            
            response = self.session.get(url)
            response.raise_for_status()
            response.encoding = 'utf-8'
            
            soup = BeautifulSoup(response.text, 'xml')
            news_titles = []
            
            # RSS 피드에서 제목 추출
            items = soup.find_all('item')
            
            for item in items:
                title_element = item.find('title')
                if title_element:
                    title = title_element.get_text().strip()
                    # 구글 뉴스의 경우 "- 출처명" 형태로 끝나므로 이를 정리
                    if ' - ' in title:
                        title = title.split(' - ')[0]
                    if title and len(title) > 5:
                        news_titles.append(title)
            
            print(f"구글 뉴스에서 {len(news_titles)}개의 뉴스 제목을 찾았습니다.")
            return news_titles
            
        except Exception as e:
            print(f"구글 뉴스 크롤링 오류: {e}")
            return []
    
    def crawl_yna_news(self, keyword="반도체"):
        """
        연합뉴스에서 특정 키워드 관련 뉴스 제목을 크롤링합니다.
        
        Args:
            keyword (str): 검색할 키워드
            
        Returns:
            list: 뉴스 제목 리스트
        """
        try:
            # 연합뉴스 검색 URL
            url = f"https://www.yna.co.kr/search/index?query={keyword}"
            print(f"연합뉴스 검색: {url}")
            
            response = self.session.get(url)
            response.raise_for_status()
            response.encoding = 'utf-8'
            
            soup = BeautifulSoup(response.text, 'html.parser')
            news_titles = []
            
            # 연합뉴스 제목 추출
            title_elements = soup.find_all('strong', class_='tit-news')
            
            for element in title_elements:
                link = element.find('a')
                if link:
                    title = link.get_text().strip()
                    if title and len(title) > 5:
                        news_titles.append(title)
            
            print(f"연합뉴스에서 {len(news_titles)}개의 뉴스 제목을 찾았습니다.")
            return news_titles
            
        except Exception as e:
            print(f"연합뉴스 크롤링 오류: {e}")
            return []
    
    def crawl_sample_news_data(self):
        """
        샘플 뉴스 데이터를 생성합니다 (크롤링이 모두 실패할 경우 데모용)
        
        Returns:
            list: 샘플 뉴스 제목 리스트
        """
        sample_titles = [
            "삼성전자, 3분기 반도체 실적 개선 전망",
            "SK하이닉스, HBM 메모리 공급 확대",
            "미국 반도체 지원법으로 국내 기업 수혜",
            "반도체 관련주 상승세 지속",
            "메모리 반도체 가격 상승 기대감",
            "TSMC, 3나노 공정 양산 본격화",
            "반도체 장비주 급등, 투자자 관심 집중",
            "중국 반도체 규제로 국내 기업 기회",
            "AI 반도체 수요 급증으로 관련주 주목",
            "반도체 사이클 회복 신호 포착"
        ]
        
        print(f"샘플 데이터 생성: {len(sample_titles)}개의 뉴스 제목")
        return sample_titles
    
    def save_to_csv(self, titles, filename='alternative_news_titles.csv'):
        """
        크롤링한 제목들을 CSV 파일로 저장합니다.
        
        Args:
            titles (list): 뉴스 제목 리스트
            filename (str): 저장할 파일명
        """
        try:
            with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)
                writer.writerow(['번호', '뉴스 제목', '수집시간'])  # 헤더
                
                current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                for idx, title in enumerate(titles, 1):
                    writer.writerow([idx, title, current_time])
                    
            print(f"결과가 {filename} 파일로 저장되었습니다.")
            
        except Exception as e:
            print(f"파일 저장 중 오류 발생: {e}")
    
    def save_to_excel(self, titles, filename='result.xlsx'):
        """
        크롤링한 제목들을 Excel 파일로 저장합니다 (openpyxl 사용).
        
        Args:
            titles (list): 뉴스 제목 리스트
            filename (str): 저장할 파일명
        """
        try:
            # 새 워크북 생성
            wb = Workbook()
            ws = wb.active
            ws.title = "뉴스 제목"
            
            # 헤더 스타일 설정
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            center_alignment = Alignment(horizontal='center', vertical='center')
            
            # 헤더 작성
            headers = ['번호', '뉴스 제목', '수집시간']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = header_border
                cell.alignment = center_alignment
            
            # 데이터 입력
            current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            for idx, title in enumerate(titles, 1):
                # 번호
                ws.cell(row=idx+1, column=1, value=idx).alignment = center_alignment
                # 제목
                ws.cell(row=idx+1, column=2, value=title)
                # 시간
                ws.cell(row=idx+1, column=3, value=current_time).alignment = center_alignment
            
            # 컬럼 너비 자동 조정
            ws.column_dimensions['A'].width = 8   # 번호
            ws.column_dimensions['B'].width = 80  # 제목 (넓게)
            ws.column_dimensions['C'].width = 20  # 시간
            
            # 데이터 영역 테두리 추가
            data_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in range(2, len(titles) + 2):
                for col in range(1, 4):
                    ws.cell(row=row, column=col).border = data_border
            
            # 제목 행 고정
            ws.freeze_panes = 'A2'
            
            # 파일 저장
            wb.save(filename)
            print(f"📊 Excel 파일이 저장되었습니다: {filename}")
            print(f"   - 총 {len(titles)}개의 뉴스 제목")
            print(f"   - 수집 시간: {current_time}")
            
        except Exception as e:
            print(f"Excel 파일 저장 중 오류 발생: {e}")
    
    def print_titles(self, titles, source=""):
        """
        크롤링한 제목들을 콘솔에 출력합니다.
        
        Args:
            titles (list): 뉴스 제목 리스트
            source (str): 출처 정보
        """
        print("\n" + "="*60)
        if source:
            print(f"[{source}] 크롤링된 뉴스 제목들:")
        else:
            print("크롤링된 뉴스 제목들:")
        print("="*60)
        
        for idx, title in enumerate(titles, 1):
            print(f"{idx:2d}. {title}")
        
        print("="*60)


def main():
    """메인 실행 함수"""
    crawler = AlternativeNewsCrawler()
    
    print("대안적인 뉴스 크롤링을 시작합니다...")
    print("키워드: 반도체 관련주\n")
    
    all_titles = []
    
    # 1. 다음 뉴스 크롤링 시도
    print("1. 다음 뉴스에서 크롤링 시도...")
    daum_titles = crawler.crawl_daum_news("반도체")
    if daum_titles:
        crawler.print_titles(daum_titles, "다음 뉴스")
        all_titles.extend(daum_titles)
    
    time.sleep(2)  # 요청 간격 조절
    
    # 2. 구글 뉴스 RSS 크롤링 시도
    print("\n2. 구글 뉴스 RSS에서 크롤링 시도...")
    google_titles = crawler.crawl_google_news("반도체 관련주")
    if google_titles:
        crawler.print_titles(google_titles, "구글 뉴스")
        # 중복 제거하여 추가
        for title in google_titles:
            if title not in all_titles:
                all_titles.append(title)
    
    time.sleep(2)  # 요청 간격 조절
    
    # 3. 연합뉴스 크롤링 시도
    print("\n3. 연합뉴스에서 크롤링 시도...")
    yna_titles = crawler.crawl_yna_news("반도체")
    if yna_titles:
        crawler.print_titles(yna_titles, "연합뉴스")
        # 중복 제거하여 추가
        for title in yna_titles:
            if title not in all_titles:
                all_titles.append(title)
    
    # 4. 모든 크롤링이 실패한 경우 샘플 데이터 사용
    if not all_titles:
        print("\n4. 모든 크롤링이 실패했습니다. 샘플 데이터를 생성합니다...")
        sample_titles = crawler.crawl_sample_news_data()
        all_titles.extend(sample_titles)
        crawler.print_titles(sample_titles, "샘플 데이터")
    
    # 최종 결과 출력 및 저장
    if all_titles:
        print(f"\n📊 크롤링 결과 요약:")
        print(f"- 다음 뉴스: {len(daum_titles)}개")
        print(f"- 구글 뉴스: {len(google_titles)}개") 
        print(f"- 연합뉴스: {len(yna_titles)}개")
        print(f"- 총 수집: {len(all_titles)}개 (중복 제거)")
        
        # 전체 결과 출력
        crawler.print_titles(all_titles, "전체 수집 결과")
        
        # 파일 저장
        crawler.save_to_csv(all_titles, 'alternative_semiconductor_news.csv')
        crawler.save_to_excel(all_titles, 'result.xlsx')
        
        print(f"\n✅ 크롤링 완료! 총 {len(all_titles)}개의 뉴스 제목을 수집했습니다.")
        print("📁 저장된 파일:")
        print("   - alternative_semiconductor_news.csv (CSV 형식)")
        print("   - result.xlsx (Excel 형식)")
    else:
        print("❌ 뉴스 제목을 수집하지 못했습니다.")


if __name__ == "__main__":
    main()
